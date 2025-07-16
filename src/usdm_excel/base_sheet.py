import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from typing import Type
from usdm_excel.cdisc_ct import CDISCCT
from usdm_excel.other_ct import OtherCT
from usdm_excel.option_manager import Options
from usdm_excel.quantity_type import QuantityType
from usdm_excel.range_type import RangeType
from usdm_excel.iso_3166 import ISO3166
from usdm_model.api_base_model import ApiBaseModelWithId
from usdm_model.quantity_range import Quantity, Range
from usdm_model.address import Address
from usdm_model.comment_annotation import CommentAnnotation
from usdm_excel.alias import Alias
from usdm_excel.option_manager import EmptyNoneOption
from usdm_excel.globals import Globals
from usdm_model.geographic_scope import GeographicScope
from usdm_model.person_name import PersonName


class BaseSheet:
    class StateError(Exception):
        pass

    class FormatError(Exception):
        pass

    def __init__(
        self,
        file_path: str,
        globals: Globals,
        sheet_name: str,
        header: int = 0,
        optional: bool = False,
        converters: dict = {},
        require: dict = {},
    ):
        self.file_path = file_path
        self.globals = globals
        self.dir_path, self.filename = os.path.split(file_path)
        self.sheet_name = sheet_name
        self.sheet = None
        self.success = False
        self._sheet_names = None
        if optional and not self._sheet_present(file_path, sheet_name):
            self._general_info(f"'{sheet_name}' not found but optional")
        else:
            if require and not self._check_cell_value(
                file_path,
                sheet_name,
                require["row"],
                require["column"],
                require["value"],
            ):
                self._general_info(
                    f"Required value {require['value']} at [{require['row']}, {require['column']}] mismatch in {sheet_name}"
                )
                pass
            else:
                self.sheet = pd.read_excel(
                    open(file_path, "rb"),
                    sheet_name=sheet_name,
                    header=header,
                    converters=converters,
                )
                self.success = True
                self._general_info("Processed sheet %s" % (sheet_name))

    def cell_empty(self, row_index, col_index):
        return pd.isnull(self.sheet.iloc[row_index, col_index])

    def column_present(self, names):
        fields = [names] if isinstance(names, str) else names
        for field in fields:
            try:
                col_index = self.sheet.columns.get_loc(field)
                return col_index
            except:
                pass
        columns = ", ".join(fields)
        raise BaseSheet.FormatError(f"Failed to detect column(s) '{columns}' in sheet")

    def read_cell_by_name(
        self, row_index, field_name, default=None, must_be_present=True
    ):
        try:
            col_index = self.column_present(field_name)
            return self.read_cell(row_index, col_index)
        except Exception as e:
            if not must_be_present:
                return ""  # if self.globals.option_manager.get(Options.EMPTY_NONE) == EmptyNoneOption.EMPTY.value else None
            elif default:
                return default
            else:
                self._error(
                    row_index,
                    -2,
                    f"Error attempting to read cell '{field_name}'. Exception: {e}",
                )
                return ""  # if self.globals.option_manager.get(Options.EMPTY_NONE) == EmptyNoneOption.EMPTY.value else None

    def read_cell(self, row_index, col_index, default=None):
        try:
            if pd.isnull(self.sheet.iloc[row_index, col_index]):
                if default:
                    return default
                else:
                    return ""  # if self.globals.option_manager.get(Options.EMPTY_NONE) == EmptyNoneOption.EMPTY.value else None
            else:
                return str(self.sheet.iloc[row_index, col_index]).strip()
        except Exception as e:
            # print(f"ROW / COL: {row_index}, {col_index}")
            self._exception(row_index, col_index, "Error reading cell", e)
            if default:
                return default
            else:
                return ""  # if self.globals.option_manager.get(Options.EMPTY_NONE) == EmptyNoneOption.EMPTY.value else None

    # Deprecate this method
    def read_cell_empty_legacy(self, row_index, col_index):
        if self.cell_empty(row_index, col_index):
            return "", True
        else:
            value = self.read_cell_empty(row_index, col_index, "-")
            if value == "":
                return "", True
            else:
                return value, False

    # Deprecate this method
    def read_cell_empty(self, row_index, col_index, empty_character):
        value = self.read_cell(row_index, col_index)
        value = "" if (value == empty_character) or (value == None) else value
        return value

    def read_cell_multiple_by_name(self, row_index, field_name, must_be_present=True):
        try:
            col_index = self.column_present(field_name)
            return self.read_cell_multiple(row_index, col_index)
        except Exception as e:
            if not must_be_present:
                return []
            else:
                self._error(
                    row_index, -2, f"Error '{e}' reading cell multiple '{field_name}'"
                )
                return []

    def read_cell_multiple(self, rindex, cindex):
        try:
            results = []
            value = self.read_cell(rindex, cindex)
            if value.strip() == "":
                return results
            for part in self._state_split(value):
                results.append(part.strip())
            return results
        except BaseSheet.StateError as e:
            self._error(rindex, cindex, f"Internal state error '{e}' reading cell")
            return []
        except BaseSheet.FormatError as e:
            self._error(
                rindex,
                cindex,
                "Format error reading cell, check the format of the cell",
            )
            return []

    def read_cell_with_previous(self, row_index, col_index, first_col_index):
        try:
            i = col_index
            while i >= first_col_index:
                if pd.isnull(self.sheet.iloc[row_index, i]):
                    i -= 1
                else:
                    return self.sheet.iloc[row_index, i].strip()
            self._warning(row_index, col_index, "No previous non-empty cell found.")
            return ""
        except Exception as e:
            self._error(
                row_index,
                col_index,
                "Error (%s) reading cell row '%s', field '%s'"
                % (e, row_index, col_index),
            )
            return ""

    def read_boolean_cell_by_name(self, row_index, field_name, must_be_present=True):
        try:
            col_index = self.column_present(field_name)
            return self.read_boolean_cell(row_index, col_index)
        except Exception as e:
            if not must_be_present:
                return False  # if self.globals.option_manager.get(Options.EMPTY_NONE) == EmptyNoneOption.EMPTY.value else None
            else:
                self._error(
                    row_index,
                    -2,
                    f"Error attempting to read cell '{field_name}'. Exception: {e}",
                )
                return False

    def read_boolean_cell(self, row_index, col_index, must_be_present=True):
        value = self.read_cell(row_index, col_index)
        if not value:
            if must_be_present:
                self._error(
                    row_index,
                    col_index,
                    "Empty cell detected where boolean value expected.",
                )
            return False
        elif value.strip().upper() in ["Y", "YES", "T", "TRUE", "1"]:
            return True
        return False

    def read_date_cell_by_name(self, row_index, field_name, must_be_present=True):
        col_index = self.column_present(field_name)
        return self.read_date_cell(row_index, col_index, must_be_present)

    def read_date_cell(self, row_index, col_index, must_be_present=True):
        cell = self.read_cell(row_index, col_index)
        try:
            # print(f"DATE: {cell}")
            return datetime.datetime.strptime(cell, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            self._exception(row_index, col_index, "Error reading date cell", e)
            return None

    def read_quantity_cell_by_name(
        self, row_index, field_name, allow_missing_units=True, allow_empty=True
    ):
        col_index = self.column_present(field_name)
        return self.read_quantity_cell(
            row_index, col_index, allow_missing_units, allow_empty
        )

    def read_quantity_cell(
        self, row_index, col_index, allow_missing_units=True, allow_empty=True
    ):
        try:
            text = self.read_cell(row_index, col_index)
            quantity = QuantityType(
                text, self.globals, allow_missing_units, allow_empty
            )
            if not quantity.errors:
                unit = Alias(self.globals).code(quantity.units_code, [])
                return (
                    None
                    if quantity.empty
                    else Quantity(
                        id=self.globals.id_manager.build_id(Quantity),
                        value=float(quantity.value),
                        unit=unit,
                    )
                )
            else:
                self._add_errors(quantity.errors, row_index, col_index)
                return None
        except Exception as e:
            self._exception(
                row_index, col_index, f"Failed to decode quantity data '{text}'", e
            )
            return None

    def read_range_cell_by_name(
        self, row_index, field_name, require_units=True, allow_empty=True
    ):
        col_index = self.column_present(field_name)
        return self.read_range_cell(row_index, col_index, require_units, allow_empty)

    def read_range_cell(
        self, row_index, col_index, require_units=True, allow_empty=True
    ):
        try:
            text = self.read_cell(row_index, col_index)
            range = RangeType(text, self.globals, require_units, allow_empty)
            if not range.errors:
                # print(f"RANGE: {range.lower} {range.upper} {range.units} {range.units_code} {range.empty} ")
                return (
                    None
                    if range.empty
                    else self.create_object(
                        Range,
                        {
                            "minValue": self.create_object(
                                Quantity,
                                {
                                    "value": float(range.lower),
                                    "unit": range.lower_units_code,
                                },
                            ),
                            "maxValue": self.create_object(
                                Quantity,
                                {
                                    "value": float(range.upper),
                                    "unit": range.upper_units_code,
                                },
                            ),
                            "isApproximate": False,
                        },
                    )
                )
            else:
                self._add_errors(range.errors, row_index, col_index)
                return None
        except Exception as e:
            self._exception(
                row_index, col_index, f"Failed to decode quantity data '{text}'", e
            )
            return None

    def read_address_cell_by_name(self, row_index, field_name, allow_empty=False):
        raw_address = self.read_cell_by_name(row_index, field_name)
        # TODO The '|' separator is preserved for legacy reasons but should be removed in the future
        if not raw_address:
            sep = ","
            parts = []
        elif "|" in raw_address:
            sep = "|"
            parts = raw_address.split(sep)
        else:
            sep = ","
            parts = self._state_split(raw_address)
        if len(parts) >= 6:
            result = self._to_address(
                self.globals.id_manager.build_id(Address),
                lines=[x.strip() for x in parts[:-5]],
                district=parts[-5].strip(),
                city=parts[-4].strip(),
                state=parts[-3].strip(),
                postal_code=parts[-2].strip(),
                country=ISO3166(self.globals).code(parts[-1].strip()),
            )
            return result
        elif allow_empty:
            return None
        else:
            col_index = self.column_present(field_name)
            self._error(
                row_index,
                col_index,
                f"Address '{raw_address}' does not contain the required fields (lines, district, city, state, postal code and country code) using '{sep}' separator characters, only {len(parts)} found",
            )
            return None

    def read_person_name_cell_by_name(self, row_index, field_name, allow_empty=False):
        raw_name = self.read_cell_by_name(row_index, field_name)
        parts = raw_name.split(",")
        if len(parts) >= 4:
            prefixes = [x.strip() for x in parts[0].strip().split(" ")]
            givenNames = [x.strip() for x in parts[1:-2]]
            familyName = parts[-2].strip()
            suffixes = [x.strip() for x in parts[-1].strip().split(" ")]
            result = self.create_object(
                cls=PersonName,
                params={
                    "text": f"{(' ').join(prefixes)}, {(', ').join(givenNames)}, {familyName}, {(' ').join(suffixes)}",
                    "prefixes": prefixes,
                    "givenNames": givenNames,
                    "familyName": familyName,
                    "suffixes": suffixes,
                },
            )
            return result
        elif allow_empty:
            return None
        else:
            col_index = self.column_present(field_name)
            self._error(
                row_index,
                col_index,
                f"Name '{raw_name}' does not contain the required fields (prefixes, given names, familt names and suffixes) using ',' separator characters, only {len(parts)} found",
            )
            return None

    def read_geographic_scopes_cell_by_name(self, row_index, field_name):
        try:
            col_index = self._get_column_index(field_name)
            return self.read_geographic_scopes_cell(row_index, col_index)
        except Exception as e:
            self._warning(
                row_index,
                -2,
                "No geographic scope column found, assuming global scope.",
            )
            return [self._scope("Global", None)]

    def read_geographic_scopes_cell(self, row_index, col_index):
        result = []
        value = self.read_cell(row_index, col_index, default="")
        if value.strip() == "":
            self._warning(
                row_index,
                col_index,
                "Empty cell detected where geographic scope values expected, assuming global scope.",
            )
            result.append(self._scope("Global", None))
        else:
            for item in self._state_split(value):
                key_value = self._key_value(
                    item, row_index, col_index, allow_single=True
                )
                if key_value[0] == "GLOBAL":
                    result.append(self._scope("Global", None))
                elif key_value[0] == "REGION":
                    code = self._country_region(key_value[1], "Region")
                    if code:
                        scope = self._scope("Region", code)
                        result.append(scope)
                elif key_value[0] == "COUNTRY":
                    code = self._country_region(key_value[1], "Country")
                    if code:
                        scope = self._scope("Country", code)
                        result.append(scope)
                else:
                    self._warning(
                        row_index,
                        col_index,
                        f"Failed to decode geographic scope '{value}'. Formats are 'Global', 'Region: <value>' or 'Country: <value>'. Assuming global scope.",
                    )
                    result.append(self._scope("Global", None))
        return result

    def _key_value(self, text: str, row_index: int, col_index: int, allow_single=False):
        if text.strip():
            parts = text.split(":")
            if len(parts) == 2:
                return [parts[0].strip().upper(), parts[1].strip()]
            elif len(parts) == 1 and allow_single:
                return [parts[0].strip().upper(), ""]
        self._error(
            row_index,
            col_index,
            f"Failed to decode key value pair '{text}', incorrect format, missing ':'?",
        )
        return ["", ""]

    def _country_region_quantity(
        self, text: str, type: str, row_index: int, col_index: int
    ):
        name_value = text.split("=")
        if len(name_value) == 2:
            quantity = self._get_quantity(name_value[1].strip())
            code = self._country_region(name_value[0].strip(), type)
            return code, quantity
        else:
            self._error(
                row_index,
                col_index,
                f"Failed to decode geographic enrollment data '{text}', incorrect format, missing '='?",
            )
            return None, None

    def _country_region(self, text: str, type: str):
        return (
            ISO3166(self.globals).region_code(text)
            if type == "Region"
            else ISO3166(self.globals).code(text)
        )

    def _get_quantity(self, text):
        quantity = QuantityType(text, self.globals, True, False)
        unit = Alias(self.globals).code(quantity.units_code, [])
        return self.create_object(
            Quantity, {"value": float(quantity.value), "unit": unit}
        )

    def _scope(self, type, code):
        scope_type = CDISCCT(self.globals).code_for_attribute(
            "GeographicScope", "type", type
        )
        alias = Alias(self.globals).code(code, []) if code else None
        return self.create_object(GeographicScope, {"type": scope_type, "code": alias})

    def _to_address(self, id, lines, city, district, state, postal_code, country):
        text = f"{(', ').join(lines)}, {city}, {district}, {state}, {postal_code}, {country.decode}"
        # text = text.replace(' ,', '')
        try:
            result = Address(
                id=id,
                text=text,
                lines=lines,
                city=city,
                district=district,
                state=state,
                postalCode=postal_code,
                country=country,
            )
        except Exception as e:
            self._general_exception(f"Failed to create Address object", e)
            result = None
        return result

    def create_object(
        self, cls: Type[ApiBaseModelWithId], params: dict, id: str = None
    ) -> object:
        try:
            params["id"] = id if id else self.globals.id_manager.build_id(cls)
            return cls(**params)
        except Exception as e:
            self._general_exception(f"Failed to create {cls.__name__} object", e)
            return None

    def add_notes(self, instance: object, note_refs: list) -> None:
        for note_ref in note_refs:
            try:
                note = self.globals.cross_references.get(CommentAnnotation, note_ref)
                if note:
                    instance.notes.append(note)
                else:
                    self._general_error(f"Failed to find note with name '{note_ref}'")
            except Exception as e:
                self._general_exception(
                    f"Failed to add note to '{object.__class__.__name__}' object", e
                )

    def read_other_code_cell_by_name(self, row_index, field_name):
        col_index = self.column_present(field_name)
        return self.read_other_code_cell(row_index, col_index)

    def read_other_code_cell(self, row_index, col_index):
        value = self.read_cell(row_index, col_index)
        if value.strip() == "":
            return None
        return self._decode_other_code(value, row_index, col_index)

    def read_other_code_cell_multiple_by_name(self, row_index, field_name):
        col_index = self.column_present(field_name)
        return self.read_other_code_cell_mutiple(row_index, col_index)

    def read_other_code_cell_mutiple(self, row_index, col_index):
        value = self.read_cell(row_index, col_index)
        result = []
        if value.strip() == "":
            return result
        for item in self._state_split(value):
            code = self._decode_other_code(item.strip(), row_index, col_index)
            if not code == None:
                result.append(code)
        return result

    def read_cdisc_klass_attribute_cell_by_name(
        self, klass, attribute, row_index, field_name, allow_empty=False
    ):
        col_index = self.column_present(field_name)
        return self.read_cdisc_klass_attribute_cell(
            klass, attribute, row_index, col_index, allow_empty
        )

    def read_cdisc_klass_attribute_cell(
        self, klass, attribute, row_index, col_index, allow_empty=False
    ):
        code = None
        value = self.read_cell(row_index, col_index)
        if value:
            code = CDISCCT(self.globals).code_for_attribute(klass, attribute, value)
            if not code:
                self._error(
                    row_index, col_index, f"CDISC CT not found for value '{value}'."
                )
        elif not allow_empty:
            self._error(
                row_index,
                col_index,
                "Empty cell detected where CDISC CT value expected.",
            )
        return code

    def read_cdisc_klass_attribute_cell_multiple_by_name(
        self, klass, attribute, row_index, field_name
    ):
        col_index = self.column_present(field_name)
        return self.read_cdisc_klass_attribute_cell_multiple(
            klass, attribute, row_index, col_index
        )

    def read_cdisc_klass_attribute_cell_multiple(
        self, klass, attribute, row_index, col_index
    ):
        result = []
        value = self.read_cell(row_index, col_index)
        if value.strip() == "":
            self._error(
                row_index,
                col_index,
                "Empty cell detected where multiple CDISC CT values expected.",
            )
            return result
        for item in self._state_split(value):
            code = CDISCCT(self.globals).code_for_attribute(
                klass, attribute, item.strip()
            )
            if code is not None:
                result.append(code)
            else:
                self._error(
                    row_index,
                    col_index,
                    f"CDISC CT not found for value '{item.strip()}'.",
                )
        return result

    def _get_cross_reference(self, klass, name):
        item = self.globals.cross_references.get(klass, name)
        if item:
            return item.id
        else:
            self._general_error(f"Unable to find {klass.__name__} with name '{name}'")
            return None

    def double_link(self, items, prev, next):
        try:
            for idx, item in enumerate(items):
                if idx == 0:
                    if (
                        self.globals.option_manager.get(Options.EMPTY_NONE)
                        == EmptyNoneOption.EMPTY.value
                    ):
                        setattr(item, prev, "")
                    else:
                        setattr(item, prev, None)
                else:
                    the_id = getattr(items[idx - 1], "id")
                    setattr(item, prev, the_id)
                if idx == len(items) - 1:
                    if (
                        self.globals.option_manager.get(Options.EMPTY_NONE)
                        == EmptyNoneOption.EMPTY.value
                    ):
                        setattr(item, next, "")
                    else:
                        setattr(item, next, None)
                else:
                    the_id = getattr(items[idx + 1], "id")
                    setattr(item, next, the_id)
        except Exception as e:
            self._general_exception(f"Error while doubly linking lists", e)

    def previous_link(self, items, prev):
        try:
            for idx, item in enumerate(items):
                if idx == 0:
                    if (
                        self.globals.option_manager.get(Options.EMPTY_NONE)
                        == EmptyNoneOption.EMPTY.value
                    ):
                        setattr(item, prev, "")
                    else:
                        setattr(item, prev, None)
                else:
                    the_id = getattr(items[idx - 1], "id")
                    setattr(item, prev, the_id)
        except Exception as e:
            self._general_exception(f"Error in previous link link {items}", e)

    def _decode_other_code(self, value, row_index, col_index):
        if value.strip() == "":
            return None
        outer_parts = value.split(":")
        if len(outer_parts) == 2:
            system = outer_parts[0].strip()
            inner_parts = outer_parts[1].strip().split("=")
            if len(inner_parts) == 2:
                version = self.globals.ct_version_manager.get(system)
                return OtherCT(self.globals).code(
                    code=inner_parts[0].strip(),
                    system=system,
                    version=version,
                    decode=inner_parts[1].strip(),
                )
            else:
                self._error(
                    row_index,
                    col_index,
                    "Failed to decode code data '%s', no '=' detected" % (value),
                )
        else:
            self._error(
                row_index,
                col_index,
                "Failed to decode code data '%s', no ':' detected" % (value),
            )
        return None

    def _to_int(self, value):
        try:
            return int(value)
        except:
            return None

    def _get_column_index(self, column_name):
        try:
            col_index = self.sheet.columns.get_loc(column_name)
            return col_index
        except Exception as e:
            pass
        raise BaseSheet.FormatError(
            f"Failed to detect column(s) '{column_name}' in sheet"
        )

    def _add_errors(self, errors, row, column):
        for error in errors:
            self._error(row, column, error)

    def _info(self, row: int, column: int | str, message: str):
        try:
            column = (
                self._get_column_index(column) if isinstance(column, str) else column
            )
            self.globals.errors_and_logging.info(
                message, self.sheet_name, row + 1, column + 1
            )
        except Exception as e:
            self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _general_info(self, message):
        self.globals.errors_and_logging.info(message, self.sheet_name)

    def _error(self, row: int, column: int | str, message: str):
        try:
            column = (
                self._get_column_index(column) if isinstance(column, str) else column
            )
            self.globals.errors_and_logging.error(
                message, self.sheet_name, row + 1, column + 1
            )
        except Exception as e:
            self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _general_error(self, message):
        self.globals.errors_and_logging.error(message, self.sheet_name)

    def _warning(self, row: int, column: int | str, message: str):
        try:
            column = (
                self._get_column_index(column) if isinstance(column, str) else column
            )
            self.globals.errors_and_logging.warning(
                message, self.sheet_name, row + 1, column + 1
            )
        except Exception as e:
            self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _general_warning(self, message):
        self.globals.errors_and_logging.warning(message, self.sheet_name)

    def _debug(self, row: int, column: int | str, message: str):
        try:
            column = (
                self._get_column_index(column) if isinstance(column, str) else column
            )
            self.globals.errors_and_logging.debug(
                message, self.sheet_name, row + 1, column + 1
            )
        except Exception as e:
            self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _general_debug(self, message):
        self.globals.errors_and_logging.debug(message, self.sheet_name)

    def _general_exception(self, message, e):
        self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _exception(self, row: int, column: int | str, message: str, e: Exception):
        try:
            column = (
                self._get_column_index(column) if isinstance(column, str) else column
            )
            self.globals.errors_and_logging.exception(
                message, e, self.sheet_name, row + 1, column + 1
            )
        except Exception as e:
            self.globals.errors_and_logging.exception(message, e, self.sheet_name)

    def _sheet_exception(self, e):
        self.globals.errors_and_logging.exception(
            f"Error [{e}] while reading sheet '{self.sheet_name}'", e, self.sheet_name
        )

    def _get_sheet_names(self, file_path):
        if not self._sheet_names:
            wb = load_workbook(file_path, read_only=True, keep_links=False)
            self._sheet_names = wb.sheetnames
        return self._sheet_names

    def _sheet_present(self, file_path, sheet_name):
        sheet_names = self._get_sheet_names(file_path)
        return sheet_name in sheet_names

    def _check_cell_value(self, file_path, sheet_name, row, column, value):
        wb = load_workbook(file_path, read_only=True, keep_links=False)
        ws = wb[sheet_name]
        # print(f"CELL={ws.cell(row, column).value}")
        return str(ws.cell(row, column).value).upper() == value

    def _state_split(self, s):
        OUT = "out"
        IN_QUOTED = "in_quoted"
        OUT_QUOTED = "out_quoted"
        IN_NORMAL = "in_normal"
        ESC = "escape"

        state = OUT
        result = []
        current = []
        exit = ""
        for c in s:
            # print(f"STATE: s: {state}, c: {c}")
            if state == OUT:
                current = []
                if c == ",":
                    result.append("")
                elif c in ['"', "'"]:
                    state = IN_QUOTED
                    exit = c
                elif c.isspace():
                    pass
                else:
                    state = IN_NORMAL
                    current.append(c)
            elif state == IN_QUOTED:
                if c == "\\":
                    state = ESC
                elif c == exit:
                    result.append("".join(current).strip())
                    state = OUT_QUOTED
                else:
                    current.append(c)
            elif state == OUT_QUOTED:
                if c == ",":
                    state = OUT
                else:
                    pass
            elif state == IN_NORMAL:
                if c == ",":
                    result.append("".join(current).strip())
                    state = OUT
                else:
                    current.append(c)
            elif state == ESC:
                if c == exit:
                    current.append(c)
                    state = IN_QUOTED
                else:
                    current.append("\\")
                    current.append(c)
            else:
                raise BaseSheet.StateError

        if state == OUT or state == OUT_QUOTED:
            pass
        elif state == IN_NORMAL:
            result.append("".join(current).strip())
        else:
            raise BaseSheet.FormatError
        return result
